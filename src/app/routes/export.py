from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from datetime import datetime
import io
import logging

from src.app.database.session import get_db
from src.app.models import Subreddit, WeeklyStats, Post, PostWeeklyStats, Comment, CommentHistory

logger = logging.getLogger(__name__)
router = APIRouter(prefix="/api/export", tags=["export"])


@router.get("/excel")
def export_excel(db: Session = Depends(get_db)):
    """
    Export all data as Excel file with multiple sheets.
    
    Returns a downloadable Excel workbook with:
    - Subreddits: subreddit summary with latest metrics
    - Weekly Stats: full weekly statistics per subreddit
    - Posts: all posts with latest metrics
    - Comments: tracked comments with latest view counts
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Define header style
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        # Sheet 1: Subreddits with latest stats
        ws1 = wb.active
        ws1.title = "Subreddits"
        headers = ["ID", "Name", "Latest Week", "Total Posts", "Impressions", "Avg Engagement Rate", "Created At"]
        ws1.append(headers)
        
        # Apply header style
        for cell in ws1[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        subreddits = db.query(Subreddit).all()
        for sub in subreddits:
            latest_stat = db.query(WeeklyStats).filter(
                WeeklyStats.subreddit_id == sub.id
            ).order_by(WeeklyStats.week.desc()).first()
            
            ws1.append([
                sub.id,
                sub.name,
                latest_stat.week if latest_stat else "N/A",
                latest_stat.total_posts if latest_stat else 0,
                latest_stat.impressions if latest_stat else 0,
                round(latest_stat.avg_engagement_rate, 2) if latest_stat else 0,
                sub.created_at.isoformat()
            ])
        
        # Sheet 2: Weekly Stats
        ws2 = wb.create_sheet("Weekly Stats")
        ws2.append(["Subreddit", "Week", "Total Posts", "Impressions", "Upvotes", "Comments", "Engagement", "Avg Engagement Rate", "Created At"])
        
        for cell in ws2[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        stats = db.query(WeeklyStats).all()
        for stat in stats:
            subreddit = db.query(Subreddit).filter(Subreddit.id == stat.subreddit_id).first()
            ws2.append([
                subreddit.name if subreddit else "",
                stat.week,
                stat.total_posts,
                stat.impressions,
                stat.upvotes,
                stat.comments,
                stat.engagement,
                round(stat.avg_engagement_rate, 2),
                stat.created_at.isoformat()
            ])
        
        # Sheet 3: Posts
        ws3 = wb.create_sheet("Posts")
        ws3.append(["ID", "Subreddit", "Title", "URL", "Author", "Created At"])
        
        for cell in ws3[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        posts = db.query(Post).all()
        for post in posts:
            subreddit = db.query(Subreddit).filter(Subreddit.id == post.subreddit_id).first()
            ws3.append([
                post.id,
                subreddit.name if subreddit else "",
                post.title,
                post.url,
                post.author,
                post.created_at.isoformat()
            ])
        
        # Sheet 4: Comments
        ws4 = wb.create_sheet("Comments")
        ws4.append(["ID", "URL", "Comment ID", "Subreddit", "Created At", "Latest View Count"])
        
        for cell in ws4[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        
        comments = db.query(Comment).all()
        for comment in comments:
            latest_history = db.query(CommentHistory).filter(
                CommentHistory.comment_id == comment.id
            ).order_by(CommentHistory.timestamp.desc()).first()
            
            ws4.append([
                comment.id,
                comment.reddit_url,
                comment.reddit_comment_id,
                comment.subreddit,
                comment.created_at.isoformat(),
                latest_history.view_count if latest_history else 0
            ])
        
        # Auto-adjust column widths
        for ws in [ws1, ws2, ws3, ws4]:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"comment_data_{datetime.utcnow().strftime('%Y-%m-%d')}.xlsx"
        logger.info(f"Excel export generated: {filename}")
        
        return FileResponse(
            io.BytesIO(output.getvalue()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            filename=filename
        )
    except ImportError:
        logger.error("openpyxl not installed")
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    except Exception as e:
        logger.error(f"Excel export failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Export failed")


@router.post("/sheets")
def sync_google_sheets(db: Session = Depends(get_db)):
    """
    Sync data to Google Sheets.
    
    Requires GOOGLE_SHEETS_ID and GOOGLE_SERVICE_ACCOUNT_JSON_PATH env variables.
    
    Returns:
    - status: "synced"
    - sheets_updated: number of sheets updated
    - rows_written: total rows written
    - timestamp: ISO8601 timestamp
    """
    try:
        import gspread
        from google.oauth2.service_account import Credentials
        from src.config import Settings
        
        settings = Settings()
        
        if not settings.google_sheets_id or not settings.google_service_account_json_path:
            logger.warning("Google Sheets config not set")
            raise HTTPException(status_code=400, detail="Google Sheets config not set")
        
        # Authenticate
        creds = Credentials.from_service_account_file(
            settings.google_service_account_json_path,
            scopes=["https://www.googleapis.com/auth/spreadsheets"]
        )
        client = gspread.authorize(creds)
        
        # Open spreadsheet
        sheet = client.open_by_key(settings.google_sheets_id)
        sheets_updated = 0
        total_rows = 0
        
        # Helper function to update/create sheet
        def update_sheet(sheet_name, headers, data_rows):
            nonlocal sheets_updated, total_rows
            try:
                ws = sheet.worksheet(sheet_name)
            except:
                ws = sheet.add_worksheet(title=sheet_name, rows=1, cols=len(headers))
            
            ws.clear()
            ws.append_row(headers)
            for row in data_rows:
                ws.append_row(row)
            
            sheets_updated += 1
            total_rows += len(data_rows)
            logger.info(f"Updated sheet '{sheet_name}' with {len(data_rows)} rows")
        
        # Sheet 1: Subreddits
        subreddit_headers = ["ID", "Name", "Latest Week", "Total Posts", "Impressions", "Avg Engagement Rate"]
        subreddit_data = []
        subreddits = db.query(Subreddit).all()
        for sub in subreddits:
            latest_stat = db.query(WeeklyStats).filter(
                WeeklyStats.subreddit_id == sub.id
            ).order_by(WeeklyStats.week.desc()).first()
            
            subreddit_data.append([
                str(sub.id),
                sub.name,
                latest_stat.week if latest_stat else "N/A",
                str(latest_stat.total_posts if latest_stat else 0),
                str(latest_stat.impressions if latest_stat else 0),
                str(round(latest_stat.avg_engagement_rate, 2) if latest_stat else 0)
            ])
        update_sheet("Subreddits", subreddit_headers, subreddit_data)
        
        # Sheet 2: Weekly Stats
        stats_headers = ["Subreddit", "Week", "Total Posts", "Impressions", "Upvotes", "Comments", "Engagement", "Avg Engagement Rate"]
        stats_data = []
        stats = db.query(WeeklyStats).all()
        for stat in stats:
            subreddit = db.query(Subreddit).filter(Subreddit.id == stat.subreddit_id).first()
            stats_data.append([
                subreddit.name if subreddit else "",
                stat.week,
                str(stat.total_posts),
                str(stat.impressions),
                str(stat.upvotes),
                str(stat.comments),
                str(stat.engagement),
                str(round(stat.avg_engagement_rate, 2))
            ])
        update_sheet("Weekly Stats", stats_headers, stats_data)
        
        # Sheet 3: Posts
        posts_headers = ["ID", "Subreddit", "Title", "URL", "Author"]
        posts_data = []
        posts = db.query(Post).all()
        for post in posts:
            subreddit = db.query(Subreddit).filter(Subreddit.id == post.subreddit_id).first()
            posts_data.append([
                str(post.id),
                subreddit.name if subreddit else "",
                post.title[:100],  # Truncate title to 100 chars
                post.url,
                post.author or ""
            ])
        update_sheet("Posts", posts_headers, posts_data)
        
        # Sheet 4: Comments
        comments_headers = ["ID", "URL", "Comment ID", "Subreddit", "Latest View Count"]
        comments_data = []
        comments = db.query(Comment).all()
        for comment in comments:
            latest_history = db.query(CommentHistory).filter(
                CommentHistory.comment_id == comment.id
            ).order_by(CommentHistory.timestamp.desc()).first()
            
            comments_data.append([
                str(comment.id),
                comment.reddit_url,
                comment.reddit_comment_id,
                comment.subreddit or "",
                str(latest_history.view_count if latest_history else 0)
            ])
        update_sheet("Comments", comments_headers, comments_data)
        
        logger.info(f"Google Sheets sync completed: {sheets_updated} sheets, {total_rows} rows")
        
        return {
            "status": "synced",
            "sheets_updated": sheets_updated,
            "rows_written": total_rows,
            "timestamp": datetime.utcnow().isoformat()
        }
    except ImportError as e:
        logger.error(f"Missing dependency: {e}")
        raise HTTPException(status_code=500, detail=f"Missing dependency: {str(e)}")
    except Exception as e:
        logger.error(f"Google Sheets sync failed: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail="Sync failed")
